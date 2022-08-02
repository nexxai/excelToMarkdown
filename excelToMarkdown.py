"""
excelToMarkdown.py
...
Usage: excelToMarkdown.py [PATH TO XLSX]

Example: excelToMarkdown.py "../Downloads/Music.xlsx"
...
Description:
Convert your multi-worksheet Excel (XLSX) spreadsheet into a block of pasteable 
Markdown (MD), perfect for dumping in a Confluence document or other 
Markdown-friendly system.
"""

import sys
from openpyxl import load_workbook

file_path = sys.argv[1]
workbook = load_workbook(file_path)

heading_count = 0
while heading_count < 1 or heading_count > 6:
    heading_count = int(input("What level of heading should each table be labeled as? [1-6] "))

heading = ""
for i in range(1, heading_count + 1):
    heading += "*"

for sheet_name in workbook.sheetnames:
    work_sheet = workbook[sheet_name]
    
    row_count = 0
    for i, row in enumerate(work_sheet):
        if row_count == 0:
            print(heading + sheet_name + heading)

        columns = []
        for column in row:
            if column.value is None:
                continue

            columns.append(column.value)

        print("|"+"|".join(columns)+"|")
        if i == 0:
            print("|"+"|".join(["---"]*len(columns))+"|")
        
        row_count += 1

